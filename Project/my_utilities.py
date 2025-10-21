# DECLARE THE FUNCTION TO SAVE DATA TO EXCEL FILES
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def flatten_list(nested_list):
    """Flatten a nested list."""
    flattened_list = []
    for sublist in nested_list:
        if isinstance(sublist, list):
            flattened_list.extend(flatten_list(sublist))
        else:
            flattened_list.append(sublist)
    return flattened_list

def save_list_to_excel(nested_list, file_path, sheet_name):
    ''' Save a nested list to an Excel file.
    If the specified sheet already exists, append the data to the next available column.
    If the sheet does not exist, create a new sheet and write the data there.'''

    # Flatten the nested list
    flattened_list = flatten_list(nested_list)

    # Load the existing workbook if it exists
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()

    # Check if the specified sheet exists in the workbook
    if sheet_name in workbook.sheetnames:
        # Get the existing sheet
        sheet = workbook[sheet_name]
        # Find the next available column
        next_column = len(list(sheet.columns)) + 1
    else:
        # Create a new sheet
        sheet = workbook.create_sheet(title=sheet_name)
        next_column = 1

    # Write data to the sheet
    for row_idx, item in enumerate(flattened_list, start=1):
        sheet.cell(row=row_idx, column=next_column, value=item)

    # Save the workbook back to the file
    workbook.save(file_path)
# ------------------------------------------------------------------------------------------------ #

# DECLARE THE FUNCTION TO READ DATA FROM EXCEL FILES
import pandas as pd
def read_excel_file(file_path):
    ''' Read an Excel file and return a dictionary of DataFrames for each sheet. '''
    xls = pd.ExcelFile(file_path)
    sheet_dict = {sheet_name: pd.read_excel(xls, sheet_name, skiprows=[1]) for sheet_name in xls.sheet_names}
    return sheet_dict
# ------------------------------------------------------------------------------------------------ #

# DECLARE THE FUNCTION TO RANDOMLY SAMPLE FROM A BETA DISTRIBUTION
'''This function is randomly sampling from a Beta distribution 
    that has been scaled and shifted to fit within a specified range [min, max].
    vector is a list or array-like structure containing three elements:
    - mean: The desired mean of the distribution.
    - max: The maximum value of the distribution.
    - min: The minimum value of the distribution.'''
import numpy as np
def beta_sampler(vector):
  mean = vector[0]
  max = vector[1]
  min = vector[2]

  alpha = max-min
  beta = alpha*(max-min)/(mean-min)-alpha
  return (max-min)*np.random.beta(alpha,beta)+min
# ------------------------------------------------------------------------------------------------ #

# Declare function to generate costs data for one year in an incremental way
'''This function generates a time series of costs for a feedstock over one year (52 weeks).
The costs are generated in an incremental manner, starting from an initial cost sampled from a Beta distribution.
The cost can fluctuate weekly based on a triangular distribution, but it is constrained within specified minimum and maximum bounds.
vector is a list or array-like structure containing three elements:
- mean: The desired mean of the distribution.
- max: The maximum value of the distribution.
- min: The minimum value of the distribution.'''
import matplotlib.pyplot as plt
def generate_cost_series(vector, number_of_points, plot=False):
  min = vector[2]
  max= vector[1]
  cost_year = np.zeros(number_of_points)
  cost_year[0] = beta_sampler(vector)
  for i in range(cost_year.shape[0]-1):
    incremented= cost_year[i]+np.random.triangular(-0.5,0,0.5)
    if incremented <=min:
      cost_year[i+1] = min
    elif incremented >=max:
      cost_year[i+1] = max
    else:
      cost_year[i+1] = incremented

  if plot:
    plt.plot(np.arange(1,len(cost_year)+1),cost_year)
  return cost_year
